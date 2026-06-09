from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import openpyxl

app = FastAPI()
FILE_NAME = "Magnets.xlsx"

class RowData(BaseModel):
    name: str
    email: str

# 1. READ DATA (GET)
@app.get("/rows")
def read_rows():
    wb = openpyxl.load_workbook(FILE_NAME)
    sheet = wb.active
    data = []
    # Assumes row 1 has headers: Name, Email
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data.append({"name": row[0], "email": row[1]})
    return data

# 2. WRITE DATA (POST)
@app.post("/rows")
def add_row(item: RowData):
    wb = openpyxl.load_workbook(FILE_NAME)
    sheet = wb.active

# 3. DELETE DATA (DELETE)
@app.delete("/rows/{name}")
def delete_row(name: str):
    wb = openpyxl.load_workbook(FILE_NAME)
    sheet = wb.active
    
    row_to_delete = None
    
    # Loop through rows to find a match (checks column 1 / "Name")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == name:
            row_to_delete = index
            break # Stop at the first match
            
    if row_to_delete is None:
        raise HTTPException(status_code=404, detail="Row not found")
        
    # Delete the row at the found index
    sheet.delete_rows(row_to_delete, amount=1)
    
    wb.save(FILE_NAME)
    return {"message": f"Row for '{name}' deleted successfully!"}

    # Append the new row to the bottom
    sheet.append([item.name, item.email])
    
    wb.save(FILE_NAME)
    return {"message": "Row added successfully!"}
